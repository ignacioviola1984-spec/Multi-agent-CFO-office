"""
adapter.py - pure, deterministic AP Control Tower export -> canonical `ap_invoices`.

This is the whole integration seam. It is a MAPPER (the same shape as
`sources/erpnext/mapper.py::map_ap_invoices`), not a `SourceConnector`: AP Control
Tower ships a PARTIAL, approved accounts-payable feed, not a full ledger, so it
must not pretend to emit P&L / balance sheet / trial balance. Instead it maps each
approved proposal row into one canonical OPEN `ap_invoices` record and merges it
into an existing canonical set, so the EXISTING engine (`finance_core`) and MCP
surface consume it with zero engine changes.

Everything here is fail-closed: a single violation rejects the file with a row +
reason (never echoing bank details or tax ids). No model, no network, no
randomness: same input -> same output.

  export rows -> validate_export (12 rules) -> map_export -> merge_into_canonical
  and build_manifest(...) emits the reproducible audit evidence alongside.
"""

import csv
import datetime
import hashlib
import json
import os
from collections import namedtuple
from decimal import Decimal, InvalidOperation

try:  # normal package import: from sources.ap_control_tower import adapter
    from .contract import (
        PAYMENT_EXPORT_COLUMNS, APPROVED_STATUS, BANK_COLUMNS, SENSITIVE_COLUMNS,
        CANONICAL_AP_COLUMNS, CANONICAL_STATUS_OPEN, TRACE_COLUMNS,
        EXPORT_TO_TRACE,
    )
except ImportError:  # repository flat-import style used by demos/tests
    from contract import (
        PAYMENT_EXPORT_COLUMNS, APPROVED_STATUS, BANK_COLUMNS, SENSITIVE_COLUMNS,
        CANONICAL_AP_COLUMNS, CANONICAL_STATUS_OPEN, TRACE_COLUMNS,
        EXPORT_TO_TRACE,
    )

Problem = namedtuple("Problem", ["row", "reason"])

# The rules the validator enforces, in evaluation order. Recorded in the manifest
# so the evidence states exactly what was checked (no free-text claims).
RULES_ENFORCED = (
    "columns_match_17_contract",
    "status_is_approved_for_proposal",
    "human_approval_evidence_present_and_valid",
    "bill_id_present",
    "vendor_present",
    "entity_explicitly_mapped",
    "currency_present_and_fx_covered",
    "amount_numeric_positive",
    "dates_valid",
    "due_not_before_issue",
    "no_duplicate_within_file",
    "no_duplicate_vs_canonical",
)


class APControlTowerError(ValueError):
    """Raised when an export cannot be ingested. Carries the list of problems
    (row + reason); its message never contains bank details or tax ids."""

    def __init__(self, problems):
        self.problems = list(problems)
        lines = [f"  row {p.row}: {p.reason}" if p.row else f"  file: {p.reason}"
                 for p in self.problems]
        super().__init__(
            "AP Control Tower export rejected ({} problem(s)):\n{}".format(
                len(self.problems), "\n".join(lines)))


# --------------------------------------------------------------------------
# Entity mapping (explicit; the export carries NO entity, so it is never inferred)
# --------------------------------------------------------------------------
class EntityMap:
    """Explicit resolution of an export row to a canonical entity_id.

    Resolution order: an exact `by_identity` mapping (the deterministic
    beneficiary + factura_documento identity) wins; then the backwards-compatible
    `by_document` mapping; otherwise a single `default` entity_id if provided;
    otherwise UNRESOLVED (the adapter rejects the row). Nothing is guessed."""

    def __init__(self, by_document=None, by_identity=None, default=None):
        self.by_document = {str(k): str(v) for k, v in (by_document or {}).items()
                            if str(v).strip()}
        self.by_identity = {str(k): str(v) for k, v in (by_identity or {}).items()
                            if str(v).strip()}
        self.default = str(default).strip() if default not in (None, "") else None

    @classmethod
    def from_dict(cls, data):
        data = data or {}
        return cls(by_document=data.get("by_document"),
                   by_identity=data.get("by_identity"),
                   default=data.get("default"))

    @classmethod
    def from_json_file(cls, path):
        with open(path, encoding="utf-8") as f:
            return cls.from_dict(json.load(f))

    def resolve(self, bill_id, vendor=None):
        bill_id = str(bill_id)
        if vendor is not None:
            identity = canonical_bill_id(vendor, bill_id)
            if identity in self.by_identity:
                return self.by_identity[identity]
        if bill_id in self.by_document:
            return self.by_document[bill_id]
        return self.default


# --------------------------------------------------------------------------
# Reading the export (CSV is the tested transport; Excel is optional)
# --------------------------------------------------------------------------
def read_export(path):
    """Read an AP Control Tower export into (columns, rows). CSV (utf-8-sig, the
    encoding AP Control Tower writes) is the primary, dependency-free path; .xlsx
    is supported via openpyxl if installed."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_export_excel(path)
    return _read_export_csv(path)


def _read_export_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        columns = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]
    return columns, rows


def _read_export_excel(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - CSV is the tested path
        raise APControlTowerError([Problem(
            0, "reading .xlsx requires the optional 'openpyxl' package; "
               "export to CSV or install openpyxl")]) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c)) for c in next(it, [])]
    rows = []
    for raw in it:
        if raw is None or all(c is None for c in raw):
            continue
        rows.append({header[i]: ("" if i >= len(raw) or raw[i] is None else str(raw[i]))
                     for i in range(len(header))})
    wb.close()
    return header, rows


# --------------------------------------------------------------------------
# Small deterministic parsers
# --------------------------------------------------------------------------
def _norm(v):
    return ("" if v is None else str(v)).strip()


def _parse_amount(s):
    """Return a Decimal, or None if not a clean number. Rejects '', thousands
    separators, and anything non-numeric (no locale guessing)."""
    s = _norm(s)
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_date(s):
    try:
        return datetime.date.fromisoformat(_norm(s))
    except ValueError:
        return None


def _parse_approval_datetime(s):
    """Return a timezone-aware ISO datetime, or None.

    AP Control Tower emits UTC offsets. Requiring an offset prevents ambiguous
    audit provenance and rejects a hand-edited status with no reliable time."""
    raw = _norm(s)
    if not raw:
        return None
    try:
        value = datetime.datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None
    if value.tzinfo is None or value.utcoffset() is None:
        return None
    return value


def _identity_part(value):
    """Case/whitespace-stable identity component; never persisted directly."""
    return " ".join(_norm(value).casefold().split())


def canonical_bill_id(vendor, source_document):
    """Collision-safe canonical id derived without retaining supplier tax data.

    Invoice numbers are unique only per supplier. Hashing normalized beneficiary
    + source document lets different suppliers reuse a number while replaying the
    same supplier/document deterministically produces the same id."""
    identity = _identity_part(vendor) + "\x1f" + _identity_part(source_document)
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:24].upper()
    return "APCT-" + digest


def _amount_str(d):
    """Canonical amount_local as a plain decimal string (no exponent)."""
    return format(d, "f")


# --------------------------------------------------------------------------
# Validation (fail-closed) - returns every problem found; never raises here
# --------------------------------------------------------------------------
def validate_export(columns, rows, entity_map, known_currencies, existing_bill_ids,
                    *, allow_due_before_issue=False, justified_docs=()):
    """Deterministically check an export against the 12 fail-closed rules.
    Returns a list of Problem(row, reason). Empty list == the file is ingestible.
    Reasons identify the row and the cause and never contain bank data or tax ids."""
    problems = []
    known = {c.upper() for c in known_currencies}
    justified = set(justified_docs)
    existing = set(existing_bill_ids)

    # Rule 1 - the 17-column contract (missing is fatal; extra is still a reject).
    missing = [c for c in PAYMENT_EXPORT_COLUMNS if c not in columns]
    extra = [c for c in columns if c not in PAYMENT_EXPORT_COLUMNS]
    if missing or extra:
        problems.append(Problem(0,
            "export columns do not match the 17-column contract"
            + (f"; missing {missing}" if missing else "")
            + (f"; unexpected {extra}" if extra else "")))
        if missing:
            return problems  # cannot map rows without the contractual columns

    seen = {}
    for i, row in enumerate(rows, start=1):
        g = lambda c: _norm(row.get(c, ""))
        bill = g("factura_documento")

        # Rule 2 - only an APPROVED proposal is ingestible.
        if g("estado") != APPROVED_STATUS:
            problems.append(Problem(i, f"estado is not '{APPROVED_STATUS}'"))

        # Rule 3 - an approved state must carry usable human-approval evidence.
        if not g("aprobado_por"):
            problems.append(Problem(i, "aprobado_por is empty; human approval evidence is required"))
        if _parse_approval_datetime(g("fecha_aprobacion")) is None:
            problems.append(Problem(i, "fecha_aprobacion is not a timezone-aware ISO datetime"))

        # Rule 4 / 5 - required identity fields present.
        if not bill:
            problems.append(Problem(i, "factura_documento (bill_id) is empty"))
        if not g("beneficiario"):
            problems.append(Problem(i, "beneficiario (vendor) is empty"))

        # Rule 6 - entity resolved ONLY from the explicit mapping.
        vendor = g("beneficiario")
        if entity_map.resolve(bill, vendor) is None:
            problems.append(Problem(i, "no explicit entity mapping for this row's "
                                       "factura_documento (entity_id is never inferred)"))

        # Rule 7 - currency present and covered by the available FX rates.
        cur = g("moneda").upper()
        if not cur:
            problems.append(Problem(i, "moneda is empty"))
        elif cur not in known:
            problems.append(Problem(i, f"currency '{cur}' is not covered by the "
                                       "available FX rates"))

        # Rule 8 - amount numeric and strictly positive.
        amt = _parse_amount(g("importe"))
        if amt is None:
            problems.append(Problem(i, "importe is empty or not numeric"))
        elif amt <= 0:
            problems.append(Problem(i, f"importe must be > 0 (got {_amount_str(amt)})"))

        # Rule 9 - valid dates.
        issue = _parse_date(g("fecha_emision"))
        due = _parse_date(g("vencimiento"))
        if issue is None:
            problems.append(Problem(i, "fecha_emision is not a valid ISO date"))
        if due is None:
            problems.append(Problem(i, "vencimiento is not a valid ISO date"))

        # Rule 10 - due date not before issue date (unless explicitly justified).
        if issue is not None and due is not None and due < issue:
            if not (allow_due_before_issue or bill in justified):
                problems.append(Problem(i, "vencimiento is before fecha_emision "
                                           "(no explicit justification provided)"))

        # Rule 11 - duplicate supplier + document identity within the file.
        identity = canonical_bill_id(vendor, bill) if vendor and bill else None
        if identity:
            if identity in seen:
                problems.append(Problem(i, "duplicate supplier + factura_documento "
                                           f"identity within the file (first seen row {seen[identity]})"))
            else:
                seen[identity] = i

        # Rule 12 - replay against the existing canonical ap_invoices.
        if identity and identity in existing:
            problems.append(Problem(i, f"supplier + factura_documento identity for '{bill}' already exists in "
                                       "canonical ap_invoices (duplicate / replay)"))

    return problems


# --------------------------------------------------------------------------
# Mapping (assumes the rows already passed validate_export)
# --------------------------------------------------------------------------
def map_export(rows, entity_map):
    """Map validated export rows -> (canonical ap_invoices rows, trace rows).

    Canonical rows carry EXACTLY the 8 contract columns; bank details and tax id
    are dropped and never persisted. Trace rows carry the richer proposal
    metadata (PO / project / method / human approval / proposal status) plus
    provenance (source_row), still WITHOUT any bank detail."""
    canonical, trace = [], []
    for i, row in enumerate(rows, start=1):
        g = lambda c: _norm(row.get(c, ""))
        bill = g("factura_documento")
        vendor = g("beneficiario")
        entity = entity_map.resolve(bill, vendor)
        canonical_id = canonical_bill_id(vendor, bill)
        amount = _amount_str(_parse_amount(g("importe")))
        currency = g("moneda").upper()

        canonical.append({
            "bill_id": canonical_id,
            "entity_id": entity,
            "vendor": vendor,
            "currency": currency,
            "amount_local": amount,
            "issue_date": g("fecha_emision"),
            "due_date": g("vencimiento"),
            "status": CANONICAL_STATUS_OPEN,   # never "paid": a proposal is not a payment
        })

        t = {"source_row": i, "source_document": bill,
             "bill_id": canonical_id, "entity_id": entity}
        for src, dst in EXPORT_TO_TRACE.items():
            t[dst] = g(src)
        t["amount_local"] = amount
        t["currency"] = currency
        trace.append({c: t.get(c, "") for c in TRACE_COLUMNS})

    return canonical, trace


# --------------------------------------------------------------------------
# Merge into the canonical set (no second engine; just appends AP rows)
# --------------------------------------------------------------------------
def merge_into_canonical(base_canonical, new_ap_rows):
    """Return a NEW canonical dict = base with `new_ap_rows` appended to
    `ap_invoices`. Every other table is carried through unchanged (copied), so
    finance_core / the MCP read the base plus the imported open payables and
    nothing else moves. Defense-in-depth: rejects a bill_id already present."""
    merged = {name: list(rows) for name, rows in base_canonical.items()}
    base_ap = list(base_canonical.get("ap_invoices", []))
    existing = {r.get("bill_id") for r in base_ap}
    collisions = sorted({r["bill_id"] for r in new_ap_rows if r["bill_id"] in existing})
    if collisions:
        raise APControlTowerError([Problem(0,
            f"refusing to merge: bill_id(s) already in canonical ap_invoices: {collisions}")])
    merged["ap_invoices"] = base_ap + [
        {c: r.get(c, "") for c in CANONICAL_AP_COLUMNS} for r in new_ap_rows]
    return merged


# --------------------------------------------------------------------------
# Evidence: a module-owned manifest (kept separate from the shared schema)
# --------------------------------------------------------------------------
def _totals(canonical_rows):
    by_ccy, by_entity = {}, {}
    for r in canonical_rows:
        ccy, ent = r["currency"], r["entity_id"]
        amt = Decimal(r["amount_local"])
        by_ccy[ccy] = by_ccy.get(ccy, Decimal(0)) + amt
        e = by_entity.setdefault(ent, {"count": 0, "by_currency": {}})
        e["count"] += 1
        e["by_currency"][ccy] = e["by_currency"].get(ccy, Decimal(0)) + amt
    by_ccy_s = {k: _amount_str(v) for k, v in sorted(by_ccy.items())}
    by_entity_s = {k: {"count": v["count"],
                       "by_currency": {c: _amount_str(a) for c, a in sorted(v["by_currency"].items())}}
                   for k, v in sorted(by_entity.items())}
    return by_ccy_s, by_entity_s


def build_manifest(canonical_rows, trace_rows, period, *, source_path=None,
                   source_bytes=None, now_iso=None):
    """Build the ingestion manifest: the reproducible audit evidence. Contains
    counts, per-currency and per-entity totals, row-level provenance, the source
    file sha256, and explicit flags that no bank data was persisted and no
    payment was executed. Contains NO bank detail or tax id."""
    by_ccy, by_entity = _totals(canonical_rows)
    if source_bytes is None and source_path is not None:
        with open(source_path, "rb") as f:
            source_bytes = f.read()
    manifest = {
        "integration": "ap_control_tower",
        "direction": "AP Control Tower export -> adapter -> canonical ap_invoices",
        "generated_at": now_iso,
        "period": period,
        "source_file": os.path.basename(source_path) if source_path else None,
        "source_sha256": hashlib.sha256(source_bytes).hexdigest() if source_bytes is not None else None,
        "imported_count": len(canonical_rows),
        "canonical_status": CANONICAL_STATUS_OPEN,
        "totals_by_currency": by_ccy,
        "by_entity": by_entity,
        "validation": {"passed": True, "rules_enforced": list(RULES_ENFORCED)},
        "provenance": [{"source_row": t["source_row"],
                        "source_document": t["source_document"],
                        "bill_id": t["bill_id"], "entity_id": t["entity_id"]}
                       for t in trace_rows],
        # Explicit, testable boundary flags:
        "bank_data_persisted": False,
        "tax_id_persisted": False,
        "payment_executed": False,
        "posted_to_ledger": False,
    }
    return manifest


# --------------------------------------------------------------------------
# Top-level ingest + output writing
# --------------------------------------------------------------------------
def ingest(base_canonical, period, *, export_path=None, columns=None, rows=None,
           entity_map=None, allow_due_before_issue=False, justified_docs=(),
           now_iso=None):
    """Read (if a path is given), validate fail-closed, map, and merge one export.

    Returns a result dict:
      {canonical_rows, trace_rows, merged_canonical, manifest,
       totals_by_currency, by_entity, imported_count}
    Raises APControlTowerError (with row + reason) on ANY violation."""
    if entity_map is None:
        raise APControlTowerError([Problem(0, "an explicit EntityMap is required "
                                              "(entity_id is never inferred)")])
    source_bytes = None
    if export_path is not None:
        with open(export_path, "rb") as f:
            source_bytes = f.read()
        columns, rows = read_export(export_path)
    if columns is None or rows is None:
        raise APControlTowerError([Problem(0, "no export provided (pass export_path "
                                              "or columns+rows)")])

    known_currencies = {"USD"} | {(_norm(r.get("currency"))).upper()
                                   for r in base_canonical.get("fx_rates", [])
                                   if _norm(r.get("period")) == period and _norm(r.get("currency"))}
    existing_bill_ids = {r.get("bill_id") for r in base_canonical.get("ap_invoices", [])}

    problems = validate_export(columns, rows, entity_map, known_currencies,
                               existing_bill_ids,
                               allow_due_before_issue=allow_due_before_issue,
                               justified_docs=justified_docs)
    if problems:
        raise APControlTowerError(problems)

    canonical_rows, trace_rows = map_export(rows, entity_map)
    merged = merge_into_canonical(base_canonical, canonical_rows)
    manifest = build_manifest(canonical_rows, trace_rows, period,
                              source_path=export_path, source_bytes=source_bytes,
                              now_iso=now_iso)
    return {
        "canonical_rows": canonical_rows,
        "trace_rows": trace_rows,
        "merged_canonical": merged,
        "manifest": manifest,
        "totals_by_currency": manifest["totals_by_currency"],
        "by_entity": manifest["by_entity"],
        "imported_count": manifest["imported_count"],
    }


def write_outputs(out_dir, result):
    """Write the module-owned evidence: trace.csv + manifest.json. Returns the
    paths. Asserts (defense-in-depth) that no sensitive column leaked into the
    trace before writing."""
    os.makedirs(out_dir, exist_ok=True)
    leaked = [c for c in SENSITIVE_COLUMNS if any(c in t for t in result["trace_rows"])]
    if leaked:  # pragma: no cover - guarded by construction
        raise APControlTowerError([Problem(0, f"internal error: sensitive fields in trace: {leaked}")])
    trace_path = os.path.join(out_dir, "trace.csv")
    with open(trace_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=TRACE_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for t in result["trace_rows"]:
            w.writerow(t)
    manifest_path = os.path.join(out_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(result["manifest"], f, indent=2, sort_keys=True)
    return {"trace": trace_path, "manifest": manifest_path}
